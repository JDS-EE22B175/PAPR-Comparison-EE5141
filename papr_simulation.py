"""
=============================================================================
PAPR Comparison: OFDMA vs F-DOSS vs IFDMA vs DFT-spread OFDMA
=============================================================================
Reference : "Generalised Multi-Carrier (GMC) ++" – K. Giridhar, IIT Madras
            Myung et al., "Single Carrier FDMA for Uplink Wireless
            Transmission," IEEE Vehicular Technology Magazine, 2006.
Course    : EE5141 – Wireless & Cellular Communications, IIT Madras
Metric    : CCDF of Peak-to-Average Power Ratio (PAPR)

Notation (from the Giridhar PDF, pages 7-11):
  N = total FFT size  (system bandwidth)
  K = number of uplink users
  P = subcarriers per user   =>  K * P = N

Key identity proven by this simulation:
  IFDMA (DFT precoding + distributed/interleaved subcarrier mapping)
  produces a time-domain signal IDENTICAL to F-DOSS (time-domain
  repetition with phase ramp). This is provable via closed-form
  analysis — see the analysis artifact for the derivation.
=============================================================================
"""

import numpy as np
import matplotlib
matplotlib.use('Agg')          # Non-interactive backend for headless runs
import matplotlib.pyplot as plt
from numpy.fft import fft, ifft
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import os

# ─────────────────────────── Global Parameters ──────────────────────────────
N = 256          # Total subcarriers (FFT size)
K = 4            # Number of uplink users
P = N // K       # Subcarriers per user  (P = 64)
NUM_ITER = 10000 # Monte-Carlo iterations
L_OS = 4         # Oversampling factor for accurate PAPR (L>=4 is standard)
QAM_ORDERS = [4, 16, 64]   # Sweep: QPSK, 16-QAM, 64-QAM

assert K * P == N, "K*P must equal N"

# ─────────────────────────── QAM Constellation ──────────────────────────────
def gen_qam_symbols(M, num_symbols):
    """
    Generate random M-QAM symbols from a square constellation.
    Symbols are normalised to unit average power.
    """
    sqrt_M = int(np.sqrt(M))
    real_vals = np.arange(-(sqrt_M - 1), sqrt_M, 2, dtype=float)
    constellation = (real_vals[:, None] + 1j * real_vals[None, :]).ravel()
    constellation /= np.sqrt(np.mean(np.abs(constellation)**2))
    indices = np.random.randint(0, M, size=num_symbols)
    return constellation[indices]

# ─────────────────────────── PAPR Calculation ───────────────────────────────
def calc_papr_dB(time_signal):
    """PAPR = max|x(n)|^2 / E[|x(n)|^2]  in dB."""
    power = np.abs(time_signal)**2
    papr_linear = np.max(power) / np.mean(power)
    return 10.0 * np.log10(papr_linear)

def oversample_freq(freq_domain_vec, N_orig, L):
    """
    Zero-pad in frequency domain to achieve L× oversampling.
    This is the standard method (ref: Myung et al. 2006, Sec. III)
    to capture inter-sample peaks for accurate PAPR measurement.
    L=4 is sufficient for convergence within 0.1 dB of continuous-time PAPR.
    """
    N_os = N_orig * L
    X_os = np.zeros(N_os, dtype=complex)
    half = N_orig // 2
    X_os[:half] = freq_domain_vec[:half]
    X_os[N_os - half:] = freq_domain_vec[half:]
    # Scale to preserve signal energy after zero-padding
    return ifft(X_os) * np.sqrt(N_os / N_orig)

# =============================================================================
# TRANSMITTER 1 : Baseline OFDMA  (PDF page 4)
# Block diagram: QAM symbols → S/P → map to P subcarriers → N-pt IFFT
# =============================================================================
def tx_ofdma(symbols_P):
    """
    Standard OFDMA: map P symbols to P CONTIGUOUS subcarriers (localized),
    then N-point IFFT. No DFT precoding. Each subcarrier independently
    carries one QAM symbol, so the time-domain signal is a superposition
    of P sinusoids → high PAPR (CLT → Gaussian-like envelope).
    """
    X = np.zeros(N, dtype=complex)
    X[0:P] = symbols_P                # Localized allocation
    return oversample_freq(X, N, L_OS)

# =============================================================================
# TRANSMITTER 2 : F-DOSS  (PDF page 8)
# Block diagram: P QAM syms → Repeat K times → phase ramp → Add CP
# NO N-point IFFT in the transmitter!
# =============================================================================
def tx_fdoss(symbols_P, user_k=0):
    """
    F-DOSS transmitter (Giridhar PDF p.8, Chang & Chen 2000):
      s(n) = d[n mod P] · exp(j·2π·k·n/N),   n = 0, ..., N-1

    For user k=0: the phase term = 1, so the output is the P QAM
    symbols tiled K times — a periodic, single-carrier-like waveform.
    This is why the PAPR is low: the envelope depends only on the
    P-symbol sequence, not on a superposition of N sinusoids.
    """
    s_time = np.tile(symbols_P, K)   # Repeat K times → length N
    n = np.arange(N)
    phase = np.exp(1j * 2.0 * np.pi * user_k * n / N)
    s_time = s_time * phase
    S_freq = fft(s_time)
    return oversample_freq(S_freq, N, L_OS)

# =============================================================================
# TRANSMITTER 3 : Interleaved OFDMA / IFDMA  (PDF page 11)
# Block diagram: P QAM syms → P×P DFT (Mixing Matrix) →
#                Place on K-spaced subcarriers → N×N IFFT → Add CP
# =============================================================================
def tx_ifdma(symbols_P, user_k=0):
    """
    IFDMA transmitter (Giridhar PDF p.11):
      1) P-pt DFT precoding  (the "Mixing Matrix")
      2) Map to DISTRIBUTED (interleaved) subcarriers: {k, K+k, 2K+k, ...}
      3) N-pt IFFT

    Mathematical identity (proven in analysis):
      For user k=0 with equi-spaced subcarriers, this reduces to
      x(n) = (1/K√P) · d[n mod P]  — identical to F-DOSS!
    """
    precoded = fft(symbols_P) / np.sqrt(P)
    X = np.zeros(N, dtype=complex)
    subcarrier_indices = user_k + np.arange(P) * K
    X[subcarrier_indices] = precoded
    return oversample_freq(X, N, L_OS)

# =============================================================================
# TRANSMITTER 4 : DFT-spread OFDMA  (SC-FDMA, LTE Uplink)
# PDF p.3: "DFT-Precoded OFDMA"; p.7: "3GPP LTE has adopted this for UL"
# Same as IFDMA but with LOCALIZED subcarrier mapping.
# =============================================================================
def tx_dft_spread_ofdma(symbols_P):
    """
    DFT-spread OFDMA / SC-FDMA / LFDMA:
      1) P-pt DFT precoding
      2) Map to P CONTIGUOUS (localized) subcarriers
      3) N-pt IFFT

    PAPR is slightly higher than IFDMA because the localized mapping
    does NOT produce a periodic time-domain structure. The signal is
    a "compressed" version of the P QAM symbols spread over N samples
    via sinc interpolation (not simple repetition).
    """
    precoded = fft(symbols_P) / np.sqrt(P)
    X = np.zeros(N, dtype=complex)
    X[0:P] = precoded                # Localized (contiguous) allocation
    return oversample_freq(X, N, L_OS)

# =============================================================================
# Monte Carlo Engine
# =============================================================================
SCHEME_NAMES = ["OFDMA", "F-DOSS", "IFDMA", "DFT-s-OFDMA"]
SCHEME_FUNCS = [tx_ofdma, tx_fdoss, tx_ifdma, tx_dft_spread_ofdma]

def run_monte_carlo(qam_order, num_iter=NUM_ITER):
    """Run Monte Carlo PAPR simulation for a given QAM modulation order."""
    num_schemes = len(SCHEME_NAMES)
    papr_db = np.zeros((num_schemes, num_iter))

    for i in range(num_iter):
        syms = gen_qam_symbols(qam_order, P)
        for s, func in enumerate(SCHEME_FUNCS):
            sig = func(syms)
            papr_db[s, i] = calc_papr_dB(sig)

    return papr_db

def compute_ccdf(papr_vals):
    """Return sorted PAPR values and corresponding CCDF probabilities."""
    sorted_papr = np.sort(papr_vals)
    ccdf = 1.0 - np.arange(1, len(sorted_papr) + 1) / len(sorted_papr)
    return sorted_papr, ccdf

def interpolate_papr_at_ccdf(papr_vals, target_ccdf):
    """Find the PAPR value at a given CCDF probability via interpolation."""
    sorted_p, ccdf = compute_ccdf(papr_vals)
    # Find the index closest to the target CCDF
    idx = np.searchsorted(-ccdf, -target_ccdf)  # decreasing ccdf
    idx = min(idx, len(ccdf) - 1)
    return sorted_p[idx]

# =============================================================================
# Main Execution
# =============================================================================
if __name__ == "__main__":
    np.random.seed(42)  # Reproducibility

    print("=" * 70)
    print("  PAPR Comparison Simulation")
    print(f"  N={N}, K={K}, P={P}, L(oversample)={L_OS}, Iterations={NUM_ITER}")
    print(f"  QAM orders: {QAM_ORDERS}")
    print("=" * 70)

    # Storage for all results
    all_results = {}  # {qam_order: papr_db array (4 x NUM_ITER)}

    for qam in QAM_ORDERS:
        print(f"\n>>> Running {qam}-QAM ...")
        t0 = time.time()
        papr_db = run_monte_carlo(qam)
        elapsed = time.time() - t0
        all_results[qam] = papr_db

        print(f"    Done in {elapsed:.1f}s")
        for s, name in enumerate(SCHEME_NAMES):
            mean_p = np.mean(papr_db[s])
            p99 = np.percentile(papr_db[s], 99)
            p999 = np.percentile(papr_db[s], 99.9)
            print(f"    {name:18s}:  mean={mean_p:.2f} dB,  "
                  f"99th%={p99:.2f} dB,  99.9th%={p999:.2f} dB")

    # ─────────────────── CCDF Plots ─────────────────────────────────────────
    colors = ["#e74c3c", "#2ecc71", "#3498db", "#9b59b6"]
    styles = ["-", "-", "--", "-."]

    # --- Individual plots per QAM order ---
    for qam in QAM_ORDERS:
        fig, ax = plt.subplots(figsize=(10, 7))
        papr_db = all_results[qam]
        for s in range(len(SCHEME_NAMES)):
            sp, ccdf = compute_ccdf(papr_db[s])
            ax.semilogy(sp, ccdf, styles[s], color=colors[s],
                        linewidth=2.0, label=SCHEME_NAMES[s])
        ax.set_xlabel("PAPR₀ (dB)", fontsize=13)
        ax.set_ylabel("Pr(PAPR > PAPR₀)  [CCDF]", fontsize=13)
        ax.set_title(f"CCDF of PAPR  |  {qam}-QAM,  N={N}, P={P}, K={K},"
                     f"  {NUM_ITER} frames", fontsize=13, fontweight='bold')
        ax.legend(fontsize=12, loc='upper right')
        ax.grid(True, which='both', alpha=0.4)
        ax.set_ylim([1e-4, 1])
        ax.set_xlim([0, 14])
        fig.tight_layout()
        fname = f"papr_ccdf_{qam}qam.png"
        fig.savefig(fname, dpi=200, bbox_inches='tight')
        print(f"  Saved: {fname}")
        plt.close(fig)

    # --- Combined plot (all QAM orders) ---
    fig, axes = plt.subplots(1, 3, figsize=(22, 7), sharey=True)
    for idx, qam in enumerate(QAM_ORDERS):
        ax = axes[idx]
        papr_db = all_results[qam]
        for s in range(len(SCHEME_NAMES)):
            sp, ccdf = compute_ccdf(papr_db[s])
            ax.semilogy(sp, ccdf, styles[s], color=colors[s],
                        linewidth=2.0, label=SCHEME_NAMES[s])
        ax.set_xlabel("PAPR₀ (dB)", fontsize=12)
        ax.set_title(f"{qam}-QAM", fontsize=13, fontweight='bold')
        ax.grid(True, which='both', alpha=0.4)
        ax.set_ylim([1e-4, 1])
        ax.set_xlim([0, 14])
        if idx == 0:
            ax.set_ylabel("Pr(PAPR > PAPR₀)  [CCDF]", fontsize=12)
        ax.legend(fontsize=10, loc='upper right')
    fig.suptitle(f"PAPR CCDF Comparison  |  N={N}, P={P}, K={K},"
                 f"  {NUM_ITER} Monte Carlo frames",
                 fontsize=15, fontweight='bold', y=1.02)
    fig.tight_layout()
    fig.savefig("papr_ccdf_combined.png", dpi=200, bbox_inches='tight')
    print("  Saved: papr_ccdf_combined.png")
    plt.close(fig)

    # ─────────────────── Excel Export ────────────────────────────────────────
    wb = Workbook()

    # --- Sheet 1: Summary table ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79",
                              fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    data_font = Font(size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["Modulation", "Scheme", "Mean PAPR (dB)",
               "PAPR @ CCDF=1%", "PAPR @ CCDF=0.1%",
               "PAPR Gain vs OFDMA @ 1% (dB)"]
    for c, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 2
    for qam in QAM_ORDERS:
        papr_db = all_results[qam]
        ofdma_1pct = interpolate_papr_at_ccdf(papr_db[0], 0.01)
        for s, name in enumerate(SCHEME_NAMES):
            mean_p = np.mean(papr_db[s])
            p_1pct = interpolate_papr_at_ccdf(papr_db[s], 0.01)
            p_01pct = interpolate_papr_at_ccdf(papr_db[s], 0.001)
            gain = ofdma_1pct - p_1pct

            ws_summary.cell(row=row, column=1,
                            value=f"{qam}-QAM").font = data_font
            ws_summary.cell(row=row, column=2, value=name).font = data_font
            ws_summary.cell(row=row, column=3,
                            value=round(mean_p, 2)).font = data_font
            ws_summary.cell(row=row, column=4,
                            value=round(p_1pct, 2)).font = data_font
            ws_summary.cell(row=row, column=5,
                            value=round(p_01pct, 2)).font = data_font
            ws_summary.cell(row=row, column=6,
                            value=round(gain, 2)).font = data_font

            for c in range(1, 7):
                ws_summary.cell(row=row, column=c).border = thin_border
                ws_summary.cell(row=row, column=c).alignment = \
                    Alignment(horizontal='center')
            row += 1

    # Auto-width columns
    for c in range(1, 7):
        ws_summary.column_dimensions[get_column_letter(c)].width = 22

    # --- Sheet 2: Comparison with "other code" results ---
    ws_compare = wb.create_sheet("Comparison vs Other Code")

    comp_headers = ["Modulation", "CCDF Target", "Metric",
                    "OFDMA", "DFT-s-OFDMA", "IFDMA", "F-DOSS",
                    "Source"]
    for c, h in enumerate(comp_headers, 1):
        cell = ws_compare.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 2
    other_results = {
        4:  {0.01: [9.70, 9.76, 9.68, 9.68],
             0.001: [10.62, 10.65, 10.56, 10.56]},
        16: {0.01: [9.76, 9.74, 9.68, 9.68],
             0.001: [10.62, 10.65, 10.57, 10.57]},
        64: {0.01: [9.70, 9.73, 9.65, 9.65],
             0.001: [10.60, 10.68, 10.60, 10.60]},
    }

    for qam in QAM_ORDERS:
        papr_db = all_results[qam]
        for ccdf_target in [0.01, 0.001]:
            # Our results
            our = [interpolate_papr_at_ccdf(papr_db[s], ccdf_target)
                   for s in range(4)]
            for c_idx, val in enumerate(our):
                ws_compare.cell(row=row, column=1,
                                value=f"{qam}-QAM").font = data_font
                ws_compare.cell(row=row, column=2,
                                value=ccdf_target).font = data_font
                ws_compare.cell(row=row, column=3,
                                value="Ours" if c_idx == 0 else "").font = data_font
            ws_compare.cell(row=row, column=3, value="Ours").font = data_font
            for c_idx in range(4):
                ws_compare.cell(row=row, column=4 + c_idx,
                                value=round(our[c_idx], 2)).font = data_font
            ws_compare.cell(row=row, column=8, value="This sim").font = data_font
            for c in range(1, 9):
                ws_compare.cell(row=row, column=c).border = thin_border
                ws_compare.cell(row=row, column=c).alignment = \
                    Alignment(horizontal='center')
            row += 1

            # Their results
            theirs = other_results[qam][ccdf_target]
            ws_compare.cell(row=row, column=1,
                            value=f"{qam}-QAM").font = data_font
            ws_compare.cell(row=row, column=2,
                            value=ccdf_target).font = data_font
            ws_compare.cell(row=row, column=3,
                            value="Other code").font = data_font
            for c_idx in range(4):
                ws_compare.cell(row=row, column=4 + c_idx,
                                value=round(theirs[c_idx], 2)).font = data_font
            ws_compare.cell(row=row, column=8,
                            value="Other code").font = data_font
            for c in range(1, 9):
                ws_compare.cell(row=row, column=c).border = thin_border
                ws_compare.cell(row=row, column=c).alignment = \
                    Alignment(horizontal='center')

            # Highlight cells where difference > 1 dB
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                   fill_type="solid")
            for c_idx in range(4):
                if abs(our[c_idx] - theirs[c_idx]) > 1.0:
                    ws_compare.cell(row=row, column=4 + c_idx).fill = red_fill
                    ws_compare.cell(row=row - 1,
                                    column=4 + c_idx).fill = red_fill
            row += 1

    for c in range(1, 9):
        ws_compare.column_dimensions[get_column_letter(c)].width = 18

    # --- Sheet 3: Raw PAPR data (first 200 iterations per scheme) ---
    ws_raw = wb.create_sheet("Raw PAPR Data (16-QAM)")
    raw_headers = ["Iteration"] + SCHEME_NAMES
    for c, h in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    papr_16 = all_results[16]
    for i in range(min(200, NUM_ITER)):
        ws_raw.cell(row=i+2, column=1, value=i+1)
        for s in range(4):
            ws_raw.cell(row=i+2, column=s+2, value=round(papr_16[s, i], 4))

    # --- Sheet 4: Simulation Parameters ---
    ws_params = wb.create_sheet("Parameters")
    params = [
        ("Parameter", "Value", "Description"),
        ("N", N, "Total FFT size (system bandwidth)"),
        ("K", K, "Number of uplink users"),
        ("P", P, "Subcarriers per user (N/K)"),
        ("L_OS", L_OS, "Oversampling factor for PAPR measurement"),
        ("NUM_ITER", NUM_ITER, "Monte Carlo iterations"),
        ("QAM_ORDERS", str(QAM_ORDERS), "Modulation orders tested"),
        ("Random Seed", 42, "For reproducibility"),
    ]
    for r, (p, v, d) in enumerate(params, 1):
        ws_params.cell(row=r, column=1, value=p).font = \
            header_font if r == 1 else data_font
        ws_params.cell(row=r, column=2, value=v).font = \
            header_font if r == 1 else data_font
        ws_params.cell(row=r, column=3, value=d).font = \
            header_font if r == 1 else data_font
        if r == 1:
            for c in range(1, 4):
                ws_params.cell(row=r, column=c).fill = header_fill
    for c in range(1, 4):
        ws_params.column_dimensions[get_column_letter(c)].width = 25

    excel_path = "papr_results.xlsx"
    wb.save(excel_path)
    print(f"\n  Excel saved: {excel_path}")

    print("\n" + "=" * 70)
    print("  All outputs generated successfully!")
    print("  Files: papr_ccdf_4qam.png, papr_ccdf_16qam.png,")
    print("         papr_ccdf_64qam.png, papr_ccdf_combined.png,")
    print(f"         {excel_path}")
    print("=" * 70)
