"""
Excel export module — writes formatted results to .xlsx.
"""

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .config import N, K, P, L_OS, NUM_ITER, QAM_ORDERS, SCHEME_NAMES
from .utils import interpolate_papr_at_ccdf


# ──────────── Styles ────────────────────────
_HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79",
                        fill_type="solid")
_HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
_DATA_FONT = Font(size=11)
_THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
               top=Side(style='thin'), bottom=Side(style='thin'))
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                        fill_type="solid")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                          fill_type="solid")


def _header_row(ws, row, headers):
    """Write a styled header row."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = _THIN


def _data_cell(ws, row, col, value):
    """Write a styled data cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _DATA_FONT
    cell.alignment = Alignment(horizontal='center')
    cell.border = _THIN
    return cell


def export_results(all_results: dict, save_path: str,
                   other_code_results: dict = None) -> None:
    """
    Export simulation results to a formatted Excel workbook.

    Parameters
    ----------
    all_results : dict {qam_order: ndarray (num_schemes, NUM_ITER)}
    save_path : str
    other_code_results : dict, optional
        {qam: {ccdf_target: [ofdma, dft, ifdma, dfdma]}} from the other code
    """
    wb = Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _header_row(ws, 1, ["Modulation", "Scheme", "Mean PAPR (dB)",
                        "PAPR @ CCDF=1%", "PAPR @ CCDF=0.1%",
                        "Gain vs OFDMA @ 1% (dB)"])

    row = 2
    for qam in QAM_ORDERS:
        papr_db = all_results[qam]
        ofdma_1pct = interpolate_papr_at_ccdf(papr_db[0], 0.01)
        for s, name in enumerate(SCHEME_NAMES):
            mean_p = np.mean(papr_db[s])
            p1 = interpolate_papr_at_ccdf(papr_db[s], 0.01)
            p01 = interpolate_papr_at_ccdf(papr_db[s], 0.001)
            gain = ofdma_1pct - p1

            _data_cell(ws, row, 1, f"{qam}-QAM")
            _data_cell(ws, row, 2, name)
            _data_cell(ws, row, 3, round(mean_p, 2))
            _data_cell(ws, row, 4, round(p1, 2))
            _data_cell(ws, row, 5, round(p01, 2))
            cell = _data_cell(ws, row, 6, round(gain, 2))
            if gain > 1.0:
                cell.fill = _GREEN_FILL
            row += 1

    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 22

    # ── Sheet 2: Comparison vs other code ───────────────────────────────
    if other_code_results:
        ws2 = wb.create_sheet("Comparison vs Other Code")
        _header_row(ws2, 1, ["Modulation", "CCDF Target",
                            "OFDMA (Ours)", "OFDMA (Other)",
                            "DFT-s (Ours)", "DFT-s (Other)",
                            "IFDMA (Ours)", "IFDMA (Other)",
                            "F-DOSS (Ours)", "DFDMA (Other)"])

        row = 2
        for qam in QAM_ORDERS:
            papr_db = all_results[qam]
            for tgt in [0.01, 0.001]:
                ours = [interpolate_papr_at_ccdf(papr_db[s], tgt)
                        for s in range(4)]
                theirs = other_code_results.get(qam, {}).get(tgt,
                                                             [0, 0, 0, 0])

                _data_cell(ws2, row, 1, f"{qam}-QAM")
                _data_cell(ws2, row, 2, tgt)

                # Pairs: ours vs theirs
                # Order: OFDMA, DFT-s, IFDMA, F-DOSS
                ours_order = [ours[0], ours[3], ours[2], ours[1]]
                for i in range(4):
                    c_ours = _data_cell(ws2, row, 3 + 2 * i,
                                        round(ours_order[i], 2))
                    c_theirs = _data_cell(ws2, row, 4 + 2 * i,
                                          round(theirs[i], 2))
                    if abs(ours_order[i] - theirs[i]) > 1.0:
                        c_ours.fill = _GREEN_FILL
                        c_theirs.fill = _RED_FILL
                row += 1

        for c in range(1, 11):
            ws2.column_dimensions[get_column_letter(c)].width = 18

    # ── Sheet 3: Raw data (16-QAM, first 200 iterations) ───────────────
    ws3 = wb.create_sheet("Raw PAPR Data (16-QAM)")
    _header_row(ws3, 1, ["Iteration"] + list(SCHEME_NAMES))
    papr_16 = all_results.get(16, all_results.get(4))
    for i in range(min(200, papr_16.shape[1])):
        _data_cell(ws3, i + 2, 1, i + 1)
        for s in range(len(SCHEME_NAMES)):
            _data_cell(ws3, i + 2, s + 2, round(float(papr_16[s, i]), 4))

    # ── Sheet 4: Parameters ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Parameters")
    _header_row(ws4, 1, ["Parameter", "Value", "Description"])
    params = [
        ("N", N, "Total FFT size"),
        ("K", K, "Number of uplink users"),
        ("P", P, "Subcarriers per user (N/K)"),
        ("L_OS", L_OS, "Oversampling factor"),
        ("NUM_ITER", NUM_ITER, "Monte Carlo iterations"),
        ("QAM_ORDERS", str(QAM_ORDERS), "Modulation orders"),
        ("RANDOM_SEED", 42, "For reproducibility"),
    ]
    for r, (p, v, d) in enumerate(params, 2):
        _data_cell(ws4, r, 1, p)
        _data_cell(ws4, r, 2, str(v))
        _data_cell(ws4, r, 3, d)
    for c in range(1, 4):
        ws4.column_dimensions[get_column_letter(c)].width = 25

    wb.save(save_path)
    print(f"  Excel saved: {save_path}")
